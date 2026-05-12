"""
Tests for ReportingAgent — Validates Excel workbook generation.
Run with: pytest tests/test_reporting_agent.py -v
"""

import os
import pytest
import asyncio
from unittest.mock import patch, MagicMock
from openpyxl import load_workbook
from agents.reporting_agent import ReportingAgent


SAMPLE_RESULTS = [
    {"module": "GL", "description": "Trial balance balanced", "severity": "FYI"},
    {"module": "AP", "description": "12 AP invoices validated", "severity": "FYI"},
    {"module": "FA", "description": "Depreciation run confirmed", "severity": "FYI"},
]


@pytest.fixture
def agent(tmp_path):
    a = ReportingAgent(period="2026-04", company="USMF")
    a.output_dir = str(tmp_path)
    return a


@pytest.fixture
def erp_data():
    return {
        "GL": MagicMock(),
        "AP": MagicMock(),
        "AR": MagicMock(),
        "FA": MagicMock(),
        "BANK": MagicMock(),
    }


@pytest.mark.asyncio
async def test_generate_report_creates_file(agent, erp_data):
    path = await agent.generate_report(erp_data, SAMPLE_RESULTS)
    assert os.path.exists(path)
    assert path.endswith("USMF_2026-04_close_report.xlsx")


@pytest.mark.asyncio
async def test_generate_report_file_contents(agent, erp_data):
    path = await agent.generate_report(erp_data, SAMPLE_RESULTS)
    wb = load_workbook(path)
    assert "Close Summary" in wb.sheetnames
    assert "Validation Results" in wb.sheetnames


@pytest.mark.asyncio
async def test_generate_report_filename_includes_period_and_company(agent, erp_data):
    path = await agent.generate_report(erp_data, SAMPLE_RESULTS)
    assert "USMF" in path
    assert "2026-04" in path
