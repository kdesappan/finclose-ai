"""
Excel Plugin — Generates formatted Excel workbooks for close reports.
Exposes workbook creation as an agent-callable tool.
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from semantic_kernel.functions import kernel_function
from typing import Annotated

TEAL = "FF00D2A9"
DARK = "FF2A3D5"
WHITE = "FFFFFF"
GREY = "FFF5F5F5"
RED = "FFFF0000"
AMBER = "FFFFBF00"
GREEN = "FF00A55"


class ExcelPlugin:
    """
    Generates formatted Excel close workbooks.
    """

    def _header(self, cell, text: str):
        cell.value = text
        cell.font = Font(bold=True, color=WHITE, name="Arial", size=10)
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _title(self, cell, text: str):
        cell.value = text
        cell.font = Font(bold=True, size=13, color=DARK, name="Arial Black")

    def _severity_color(self, severity: str) -> str:
        return {"BLOCK": RED, "REVIEW": AMBER, "FYI": GREEN}.get(severity, "FF000000")

    @kernel_function(description="Generate the close summary Excel workbook from validation results.")
    def generate_close_workbook(
        self,
        period: Annotated[str, "Accounting period, e.g. 2026-04"],
        company: Annotated[str, "D365 company ID, e.g. USMF"],
        validation_json: Annotated[str, "JSON string of validation results list"],
        output_dir: Annotated[str, "Directory path to save the workbook"] = "reports",
    ) -> str:
        import json
        results = json.loads(validation_json)
        os.makedirs(output_dir, exist_ok=True)

        wb = Workbook()
        self._build_summary(wb, period, company, results)
        self._build_blockers(wb, results)
        self._build_full_checklist(wb, results)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        path = os.path.join(output_dir, f"{company}_{period}_close_report.xlsx")
        wb.save(path)
        return path

    def _build_summary(self, wb: Workbook, period: str, company: str, results: list):
        ws = wb.create_sheet("Close Summary")
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 12

        self._title(ws["A1"], f"FinClose AI — {company} | Period: {period}")
        ws.merge_cells("A1:B1")
        ws.row_dimensions[1].height = 28

        blocks = sum(1 for r in results if r["severity"] == "BLOCK")
        reviews = sum(1 for r in results if r["severity"] == "REVIEW")
        fyis = sum(1 for r in results if r["severity"] == "FYI")
        total = len(results)

        data = [
            ("Total Checks Run", total),
            ("Blockers (BLOCK)", blocks),
            ("Under Review (REVIEW)", reviews),
            ("Informational (FYI)", fyis),
            ("Close Status", "READY" if blocks == 0 else "BLOCKED"),
        ]
        for row, (label, value) in enumerate(data, 3):
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = Font(bold=True, name="Arial")
            cell = ws.cell(row=row, column=2)
            cell.value = value
            if label == "Close Status":
                cell.font = Font(bold=True, color=GREEN if blocks == 0 else RED)

    def _build_blockers(self, wb: Workbook, results: list):
        ws = wb.create_sheet("Blockers")
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 14
        headers = ["Module", "Description", "Severity"]
        for col, h in enumerate(headers, 1):
            self._header(ws.cell(row=1, column=col), h)
        blockers = [r for r in results if r["severity"] == "BLOCK"]
        for row, r in enumerate(blockers, 2):
            ws.cell(row=row, column=1).value = r.get("module", "")
            ws.cell(row=row, column=2).value = r.get("description", "")
            cell = ws.cell(row=row, column=3)
            cell.value = r.get("severity", "")
            cell.font = Font(bold=True, color=self._severity_color(r.get("severity", "")))

    def _build_full_checklist(self, wb: Workbook, results: list):
        ws = wb.create_sheet("Full Checklist")
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 14
        headers = ["Module", "Description", "Severity"]
        for col, h in enumerate(headers, 1):
            self._header(ws.cell(row=1, column=col), h)
        for row, r in enumerate(results, 2):
            ws.cell(row=row, column=1).value = r.get("module", "")
            ws.cell(row=row, column=2).value = r.get("description", "")
            cell = ws.cell(row=row, column=3)
            cell.value = r.get("severity", "")
            cell.font = Font(bold=True, color=self._severity_color(r.get("severity", "")))
            if row % 2 == 0:
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=GREY)
